#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pysvn
import time
import sys
import re
import requests
import json

try:
	from openpyxl import Workbook
	excel_module = 'openpyxl'
except ImportError:
	import xlwt
	excel_module = 'xlwt' 

#配置段
root_path = "https://192.168.1.25:6666/svn/AntMan/branches/publish/20210923_3_55_0_week48" #目标地址，结尾不要有/
start = 178356 #起始svn版本号(含)
end = None #结束svn版本号（含），以最新的为基准请填None
#design_path = '/client/'
design_path = '/design/X_Data/'

nowaday = time.localtime(time.time())
if excel_module == 'openpyxl':
	ext = '.xlsx'
elif excel_module == 'xlwt':
	ext = '.xls'
excel_filename = 'checklist' + str(nowaday.tm_mon) + '.' + str(nowaday.tm_mday) + ext

if excel_module == 'openpyxl':
	wb = Workbook()
	ws = wb.active
	ws.title = 'checklist'
	ws['A1'] = '确认人'
	ws['B1'] = 'Revision'
	ws['C1'] = 'Author'
	ws['D1'] = 'Date'
	ws['E1'] = 'Message'
	ws['F1'] = '单号'
	ws['G1'] = '测试负责人'
	ws['H1'] = '版本'
	ws['I1'] = 'Files'
	idx = 2
elif excel_module == 'xlwt':
	wb = xlwt.Workbook(encoding = 'utf-8')
	ws = wb.add_sheet(u'checklist')
	navi = ['Revision', 'Author', 'Date', 'Message', 'Files']
	for i in range(0, len(navi)):
		ws.write(0, i, navi[i])
	idx = 1

tester = {158:"黄洋",63:"蒋昊",88:"蒋骁瑜",60:"周仁佳",152:"张肇哲",39:"朱红",161:"郑涛",163:"单号不存在",170:"成澳",179:"甘杭波",180:"徐佳伟"}
def getredminedate(redminenum):
	# redminenum = redminenum
	print(redminenum)
	url = 'http://192.168.1.25:7777/redmine/issues/'+redminenum
	r = requests.get(url+".json",auth=('huangyang', 'hy123456'))
	try:
		iss = r.json()["issue"]
	except:
		iss = {"fixed_version":{
            "id":419,
            "name":"提交单号不存在"
        },"custom_fields":[
            {
                "id":18,
                "name":"外放版本",
                "value":"4.23 全球版本"
            },
            {
                "id":22,
                "name":"地区开关",
                "value":[
                    "大陆CN 外放 O",
                    "港台TW 外放 O",
                    "日本JP 外放 O"
                ]
            },
            {
                "id":20,
                "name":"任务属性",
                "value":"策划 开发"
            },
            {
                "id":4,
                "name":"测试负责人",
                "value":"163"
            }]}
		url = "单不存在"
	# print(iss)
	time.sleep(0.2)
	try:
		return [iss["fixed_version"]["name"],tester[int(iss["custom_fields"][3]["value"])],url]
	except KeyError:
		return ["警告","缺少相关字段",url]

class svnLog:
	def setRevision(self, revision):
		self.revision = revision.number

	def setAuthor(self, author):
		self.author = author

	def setDate(self, date):
		self.date = time.strftime(('%Y-%m-%d %X'), time.gmtime(date+28800))

	def setMessage(self, message):
		self.message = message
		red = re.compile(r'#(\d{5})')
		self.redminenums = red.findall(message)
		if self.redminenums:
			try:	
				redminedata = getredminedate(self.redminenums[0])
				self.tester = redminedata[1]
				self.versionsnum = redminedata[0]	
				self.hyperlink = redminedata[2]
			except IndexError:
				self.redminenums = []
				self.tester = "单异常"
				self.versionsnum = "单异常"
		else:
			self.tester = "无单号"
			self.versionsnum = "无单号"
	def setFiles(self, files):
		self.files = files


		

	def writeExcel(self):
		global idx
		if excel_module == 'openpyxl':
			file_col = 0
			ws['A'+str(idx)] = self.revision
			ws['B'+str(idx)] = self.author
			ws['C'+str(idx)] = self.date
			ws['D'+str(idx)] = self.message
			if self.redminenums:
				ws['D'+str(idx)].hyperlink = self.hyperlink
				ws['E'+str(idx)] = self.redminenums[0]
			else:
				ws['E'+str(idx)] = "无单号"
			ws['F'+str(idx)] = self.tester
			ws['G'+str(idx)] = self.versionsnum
			for file in self.files:
				try:
					ws[chr(ord('H')+file_col)+str(idx)] = file
				except:
					print(file)
				file_col += 1
				if file_col > 20:
					file_col = 0
					idx += 1
			self.clear()
			idx += 1
		elif excel_module == 'xlwt':
			file_col = 0
			ws.write(idx, 0, self.revision)
			ws.write(idx, 1, self.author)
			ws.write(idx, 2, self.date)
			ws.write(idx, 3, self.message)
			for file in self.files:
				ws.write(idx, 4 + file_col, file)
				file_col += 1
			self.clear()
			idx += 1

	def clear(self):
		self.revision = ''
		self.author = ''
		self.date = ''
		self.message = ''
		self.files = ''

tmpSvnLog = svnLog()

client = pysvn.Client()
start_rev = pysvn.Revision(pysvn.opt_revision_kind.number, start)
if end == None:
	end_rev = pysvn.Revision(pysvn.opt_revision_kind.head)
else:
	end_rev = pysvn.Revision(pysvn.opt_revision_kind.number, end)

new_path = root_path + design_path
logs = client.log(new_path, end_rev, start_rev, True)

for log in logs:
	tmpSvnLog = svnLog()
	files = []
	tmpSvnLog.setRevision(log.revision)
	tmpSvnLog.setAuthor(log.author)
	tmpSvnLog.setDate(log.date)
	tmpSvnLog.setMessage(log.message)
	for mod_files in log.changed_paths:
		files.append(mod_files.path.partition(design_path)[2])
	tmpSvnLog.setFiles(files)
	tmpSvnLog.writeExcel()

if excel_module == 'openpyxl':
	try:
		wb.save(filename = excel_filename)
	except:
		wb.save(filename = "tmp.xlsx")	
elif excel_module == 'xlwt':
	wb.save(excel_filename)
