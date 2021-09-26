from datetime import date
import pysvn
import time
import openpyxl
import re
import requests

timing = time.strftime(r"%m.%d", time.localtime()) 
timing1 = str(timing)
timing2 = timing1[1:5]
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = 'client'
ws2 = wb.create_sheet('X_Data')
ws3 = wb.create_sheet('Y_Tools')

ws1.cell(1,1).value = '确认人'
ws2.cell(1,1).value = '确认人'
ws3.cell(1,1).value = '确认人'
ws1.cell(1,2).value = 'revision'
ws2.cell(1,2).value = 'revision'
ws3.cell(1,2).value = 'revision'
ws1.cell(1,3).value = 'author'
ws2.cell(1,3).value = 'author'
ws3.cell(1,3).value = 'author'
ws1.cell(1,4).value = 'date'
ws2.cell(1,4).value = 'date'
ws3.cell(1,4).value = 'date'
ws1.cell(1,5).value = 'message'
ws2.cell(1,5).value = 'message'
ws3.cell(1,5).value = 'message'
ws1.cell(1,6).value = '单号'
ws2.cell(1,6).value = '单号'
ws3.cell(1,6).value = '单号'
ws1.cell(1,7).value = '测试负责人'
ws2.cell(1,7).value = '测试负责人'
ws3.cell(1,7).value = '测试负责人'
ws1.cell(1,8).value = 'files'
ws2.cell(1,8).value = 'files'
ws3.cell(1,8).value = 'files'

root_path = "https://192.168.1.25:6666/svn/AntMan/trunk" #目标地址，结尾不要有/
start = 175314 #起始svn版本号(含)
end = 178302 #结束svn版本号（含），以最新的为基准请填None

design_path = '/client/'
design_path2 = '/design/X_Data/'
design_path3 = '/design/Y_Tools/'

client = pysvn.Client()

start_rev = pysvn.Revision(pysvn.opt_revision_kind.number, start)
if end == None:
	end_rev = pysvn.Revision(pysvn.opt_revision_kind.head)
else:
	end_rev = pysvn.Revision(pysvn.opt_revision_kind.number, end)


new_path = root_path + design_path
logs = client.log(new_path, end_rev, start_rev, True)


new_path2 = root_path + design_path2
logs2 = client.log(new_path2, end_rev, start_rev, True)

new_path3 = root_path + design_path3
logs3 = client.log(new_path3, end_rev, start_rev, True)



class svnLog:
	def getlog(self, revision,author,date,message,file):			###获取log信息给对象
		global count
		self.revision = revision.number
		
		self.author = author
		
		timeArray = time.localtime(date)
		otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
		self.date = otherStyleTime
		
		self.message = message
		
		a = re.search(r'CHECK',message)
		if a:
			list1 = re.findall(r'\d{5}',message)
			self.number = list1[1]
		else:
			a = re.search(r'\d{5}',message)
			if a:
				self.number = a.group()
			else:
				self.number = '无单号'
		
		self.tester = 'tester'
		testers = {158:"黄洋",63:"蒋昊",88:"蒋骁瑜",120:"警告",60:"周仁佳",152:"张肇哲",39:"朱红",161:"郑涛",163:"单号不存在",170:"成澳",179:"甘杭波",180:"徐佳伟",200:'警告',114:'警告',108:'警告',103:'警告',101:'警告'}
		url = 'http://192.168.1.25:7777/redmine/issues/'
		r = requests.get(url+str(self.number)+'.json')
		try:
			iss = r.json()["issue"]["custom_fields"][3]
		except:
			iss = {'value':200}
		self.tester = testers[int(iss["value"])]
		self.file = file
		
		
	
	def write1(self):			###写client表
		global count
		
		ws1['B'+str(count)] = self.revision
		ws1['C'+str(count)] = self.author
		ws1['D'+str(count)] = self.date
		ws1['E'+str(count)] = self.message.replace("\n"," ")

		ws1['F'+str(count)] = self.number
		ws1['G'+str(count)] = self.tester
		file_col2 = 0
		for file in self.file:
			
			try:
				ws1[chr(ord('H')+file_col2)+str(count)] = file
			except:
				pass
			file_col2 += 1
			if file_col2 > 20:
				file_col2 = 0
		
		count = count + 1

	def write2(self):			###写X_Data表
		global count
		
		ws2['B'+str(count)] = self.revision
		ws2['C'+str(count)] = self.author
		ws2['D'+str(count)] = self.date
		ws2['E'+str(count)] = self.message.replace("\n",' ')

		ws2['F'+str(count)] = self.number
		ws2['G'+str(count)] = self.tester
		file_col = 0
		for file in self.file:
			
			try:
				ws2[chr(ord('H')+file_col)+str(count)] = file
			except:
				pass
			file_col += 1
			if file_col > 20:
				file_col = 0
	
		count = count + 1

	def write3(self):
		global count
		
		ws3['B'+str(count)] = self.revision
		ws3['C'+str(count)] = self.author
		ws3['D'+str(count)] = self.date
		ws3['E'+str(count)] = self.message.replace("\n",' ')

		ws3['F'+str(count)] = self.number
		ws3['G'+str(count)] = self.tester
		file_col = 0
		for file in self.file:
			
			try:
				ws3[chr(ord('H')+file_col)+str(count)] = file
			except:
				pass
			file_col += 1
			if file_col > 20:
				file_col = 0
	
		count = count + 1



checklist = svnLog()			###完成client
count = 2
for i in logs:
	file2 = []
	for mod_files in i.changed_paths:
		file2.append(mod_files.path.partition(design_path)[2])
	checklist.getlog(i.revision,i.author,i.date,i.message,file2)
	checklist.write1()
	
	



checklist2 = svnLog()			###完成X_Data
count = 2
for i in logs2:
	file = []
	for mod_files in i.changed_paths:
		file.append(mod_files.path.partition(design_path2)[2])
	
	checklist2.getlog(i.revision,i.author,i.date,i.message,file)
	checklist2.write2()


checklist3 = svnLog()			###完成Y_Tools
count = 2
for i in logs3:
	file3 = []
	for mod_files in i.changed_paths:
		file3.append(mod_files.path.partition(design_path3)[2])
	
	checklist3.getlog(i.revision,i.author,i.date,i.message,file3)
	checklist3.write3()




j = 0							#####删除client自动导表内容
list1 = []
max = ws1.max_row+1
for i in range(1,max,1):		
	if re.search('Flow',str(ws1.cell(i,3).value)) or re.search('Server',str(ws1.cell(i,3).value)):
		list1.append(i)
for i in list1:
	k=i-j
	ws1.delete_rows(k)
	j=j+1


j = 0							#####删除design自动导表内容
list1 = []
max = ws2.max_row+1
for i in range(1,max,1):		
	if re.search('Flow',str(ws2.cell(i,3).value)):
		list1.append(i)
for i in list1:
	k=i-j
	ws2.delete_rows(k)
	j=j+1

ws1.column_dimensions['C'].width = 15.0    ###调整列宽
ws1.column_dimensions['D'].width = 21.0
ws1.column_dimensions['E'].width = 80.0
ws2.column_dimensions['C'].width = 15.0
ws2.column_dimensions['D'].width = 21.0
ws2.column_dimensions['E'].width = 80.0
ws3.column_dimensions['C'].width = 15.0
ws3.column_dimensions['D'].width = 21.0
ws3.column_dimensions['E'].width = 80.0




wb.save(f'C:/Users/A/Downloads/checklist/{timing2}checklist.xlsx')